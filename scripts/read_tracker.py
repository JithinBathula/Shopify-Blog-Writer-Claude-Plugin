#!/usr/bin/env python3
"""
Read and update the blog tracker Excel file.

Usage:
    # Get the next eligible row (by publishing priority, then row order)
    python read_tracker.py --tracker /path/to/blog-tracker.xlsx --action next

    # Mark a row as in-progress
    python read_tracker.py --tracker /path/to/blog-tracker.xlsx --action set-status --row 30 --status in-progress

    # Mark a row as done with Shopify details
    python read_tracker.py --tracker /path/to/blog-tracker.xlsx --action update \
        --row 30 --status done --shopify-id 662924853496 --shopify-url collagen-for-bone-health

    # Count remaining pending/planned rows
    python read_tracker.py --tracker /path/to/blog-tracker.xlsx --action count

Returns JSON to stdout for easy parsing by Claude.
"""

import argparse
import json
import sys
from datetime import date
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print(json.dumps({"error": "openpyxl not installed. Run: pip install openpyxl --break-system-packages"}))
    sys.exit(1)


# All known column names mapped to their header text
COLUMN_MAP = {
    "#": "#",
    "Title": "Title",
    "Author": "Author",
    "Length": "Length",
    "Status": "Status",
    "Shopify ID": "Shopify ID",
    "Shopify URL": "Shopify URL",
    "Completed Date": "Completed Date",
    "Format": "Format",
    "Target Keyword": "Target Keyword",
    "Gap Type": "Gap Type",
    "Strategic Rationale": "Strategic Rationale",
    "Hidden Intent": "Hidden Intent",
    "Key Arguments": "Key Arguments",
    "Product Tie-In": "Product Tie-In",
    "Recommended Word Count": "Recommended Word Count",
    "Recommended Structure": "Recommended Structure",
    "Publishing Priority": "Publishing Priority",
}

ELIGIBLE_STATUSES = {"pending", "planned"}


def discover_columns(ws):
    """Read header row and return {column_name: column_index} mapping."""
    headers = {}
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col_idx).value
        if val and str(val).strip():
            headers[str(val).strip()] = col_idx
    return headers


def row_to_dict(ws, row_idx, headers):
    """Extract all column values for a given row as a dict."""
    data = {}
    for col_name, col_idx in headers.items():
        cell_val = ws.cell(row=row_idx, column=col_idx).value
        # Convert to appropriate Python type
        if cell_val is None:
            data[col_name] = None
        elif isinstance(cell_val, (int, float)):
            data[col_name] = cell_val
        else:
            data[col_name] = str(cell_val).strip()
    return data


def find_eligible_rows(ws, headers):
    """Find all rows with status = pending or planned. Returns list of (row_idx, row_data)."""
    status_col = headers.get("Status")
    if not status_col:
        return []

    eligible = []
    for row_idx in range(2, ws.max_row + 1):
        status_val = ws.cell(row=row_idx, column=status_col).value
        if status_val and str(status_val).strip().lower() in ELIGIBLE_STATUSES:
            row_data = row_to_dict(ws, row_idx, headers)
            eligible.append((row_idx, row_data))
    return eligible


def select_next_row(eligible_rows):
    """Pick the next row to process: lowest Publishing Priority first, then row order."""
    if not eligible_rows:
        return None

    def sort_key(item):
        row_idx, data = item
        priority = data.get("Publishing Priority")
        if priority is not None:
            try:
                return (0, float(priority), row_idx)
            except (ValueError, TypeError):
                pass
        return (1, 0, row_idx)  # No priority → sort after prioritized rows

    eligible_rows.sort(key=sort_key)
    return eligible_rows[0]


def action_next(tracker_path):
    """Find and return the next eligible row."""
    wb = load_workbook(tracker_path)
    ws = wb.active
    headers = discover_columns(ws)
    eligible = find_eligible_rows(ws, headers)

    if not eligible:
        return {"found": False, "message": "No pending or planned rows in the tracker.", "remaining": 0}

    row_idx, row_data = select_next_row(eligible)

    # Build a clean output with snake_case keys
    result = {
        "found": True,
        "row_index": row_idx,
        "row_number": row_data.get("#"),
        "title": row_data.get("Title"),
        "author": row_data.get("Author"),
        "length": row_data.get("Length"),
        "status": row_data.get("Status"),
        "format": row_data.get("Format"),
        "target_keyword": row_data.get("Target Keyword"),
        "gap_type": row_data.get("Gap Type"),
        "strategic_rationale": row_data.get("Strategic Rationale"),
        "hidden_intent": row_data.get("Hidden Intent"),
        "key_arguments": row_data.get("Key Arguments"),
        "product_tie_in": row_data.get("Product Tie-In"),
        "recommended_word_count": row_data.get("Recommended Word Count"),
        "recommended_structure": row_data.get("Recommended Structure"),
        "publishing_priority": row_data.get("Publishing Priority"),
        "remaining": len(eligible),
    }
    wb.close()
    return result


def action_set_status(tracker_path, row_idx, new_status):
    """Set the status of a specific row."""
    wb = load_workbook(tracker_path)
    ws = wb.active
    headers = discover_columns(ws)

    status_col = headers.get("Status")
    if not status_col:
        return {"error": "Status column not found in tracker."}

    ws.cell(row=row_idx, column=status_col, value=new_status)
    wb.save(tracker_path)
    wb.close()
    return {"success": True, "row": row_idx, "status": new_status}


def action_update(tracker_path, row_idx, status, shopify_id=None, shopify_url=None):
    """Update a row with completion data."""
    wb = load_workbook(tracker_path)
    ws = wb.active
    headers = discover_columns(ws)

    if "Status" in headers:
        ws.cell(row=row_idx, column=headers["Status"], value=status)
    if shopify_id and "Shopify ID" in headers:
        ws.cell(row=row_idx, column=headers["Shopify ID"], value=shopify_id)
    if shopify_url and "Shopify URL" in headers:
        ws.cell(row=row_idx, column=headers["Shopify URL"], value=shopify_url)
    if "Completed Date" in headers and status == "done":
        ws.cell(row=row_idx, column=headers["Completed Date"], value=date.today().isoformat())

    wb.save(tracker_path)
    wb.close()
    return {"success": True, "row": row_idx, "status": status}


def action_count(tracker_path):
    """Count remaining eligible rows."""
    wb = load_workbook(tracker_path)
    ws = wb.active
    headers = discover_columns(ws)
    eligible = find_eligible_rows(ws, headers)
    wb.close()
    return {"remaining": len(eligible)}


def main():
    parser = argparse.ArgumentParser(description="Read and update the blog tracker Excel file.")
    parser.add_argument("--tracker", required=True, help="Path to blog-tracker.xlsx")
    parser.add_argument("--action", required=True, choices=["next", "set-status", "update", "count"])
    parser.add_argument("--row", type=int, help="Row index (Excel row number, 1-based, header is row 1)")
    parser.add_argument("--status", help="New status value")
    parser.add_argument("--shopify-id", help="Shopify article ID")
    parser.add_argument("--shopify-url", help="Shopify URL handle")

    args = parser.parse_args()

    tracker = Path(args.tracker)
    if not tracker.exists():
        print(json.dumps({"error": f"Tracker file not found: {tracker}"}))
        sys.exit(1)

    if args.action == "next":
        result = action_next(tracker)
    elif args.action == "set-status":
        if not args.row or not args.status:
            print(json.dumps({"error": "--row and --status are required for set-status"}))
            sys.exit(1)
        result = action_set_status(tracker, args.row, args.status)
    elif args.action == "update":
        if not args.row or not args.status:
            print(json.dumps({"error": "--row and --status are required for update"}))
            sys.exit(1)
        result = action_update(tracker, args.row, args.status, args.shopify_id, args.shopify_url)
    elif args.action == "count":
        result = action_count(tracker)
    else:
        result = {"error": f"Unknown action: {args.action}"}

    print(json.dumps(result, default=str))


if __name__ == "__main__":
    main()
